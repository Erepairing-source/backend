"""
Organization Admin endpoints - Comprehensive management
"""
from fastapi import APIRouter, Depends, HTTPException, status, Body, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from sqlalchemy import func, and_, or_
from sqlalchemy.orm import joinedload
from datetime import datetime, timedelta, timezone
import json
import io
import traceback

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.core.database import get_db
from app.core.permissions import require_role
from app.core.data_isolation import check_organization_access, enforce_organization_isolation
from app.models.user import User, UserRole
from app.models.organization import Organization, OrganizationHierarchy
from app.models.product import Product, ProductModel, ProductCategory
from app.models.sla_policy import SLAPolicy, ServicePolicy, SLAType
from app.models.integration import Integration, IntegrationType, IntegrationStatus
from app.models.ticket import Ticket, TicketStatus
from app.models.device import Device
from app.models.inventory import Part, Inventory, InventoryTransaction, ReorderRequest
from app.models.location import Country, State, City
from app.models.subscription import Subscription, Plan, BillingPeriod
from app.models.product_part import ProductPart
from app.services.ai.route_optimization import RouteOptimizationService

router = APIRouter()
route_optimizer = RouteOptimizationService()


@router.get("/dashboard")
async def get_org_admin_dashboard(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get comprehensive organization admin dashboard data"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    org_id = current_user.organization_id
    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    
    try:
        # Get comprehensive stats
        total_tickets = db.query(func.count(Ticket.id)).filter(Ticket.organization_id == org_id).scalar() or 0
        
        # Handle ticket status enum properly
        open_tickets = db.query(func.count(Ticket.id)).filter(
            Ticket.organization_id == org_id,
            Ticket.status.in_([TicketStatus.CREATED, TicketStatus.ASSIGNED, TicketStatus.IN_PROGRESS])
        ).scalar() or 0
        
        total_users = db.query(func.count(User.id)).filter(User.organization_id == org_id).scalar() or 0
        
        # Get devices count (handle if table doesn't exist)
        try:
            total_devices = db.query(func.count(Device.id)).filter(Device.organization_id == org_id).scalar() or 0
        except Exception:
            total_devices = 0
        
        # Get products count (handle if table doesn't exist)
        try:
            total_products = db.query(func.count(Product.id)).filter(Product.organization_id == org_id).scalar() or 0
        except Exception:
            total_products = 0
        
        # Get SLA policies count (handle if table doesn't exist)
        try:
            total_sla_policies = db.query(func.count(SLAPolicy.id)).filter(SLAPolicy.organization_id == org_id).scalar() or 0
        except Exception:
            total_sla_policies = 0
        
        # Get integrations count (handle if table doesn't exist)
        try:
            active_integrations = db.query(func.count(Integration.id)).filter(
                Integration.organization_id == org_id,
                Integration.is_active == True
            ).scalar() or 0
        except Exception:
            active_integrations = 0
        
        # Get subscription with plan relationship loaded
        subscription = db.query(Subscription).options(
            joinedload(Subscription.plan)
        ).filter(Subscription.organization_id == org_id).first()
        
        subscription_data = None
        if subscription:
            subscription_data = {
                "plan_name": subscription.plan.name if subscription.plan else None,
                "status": str(subscription.status) if subscription.status else None,
                "end_date": subscription.end_date.isoformat() if subscription.end_date else None
            }
        
        return {
            "organization": {
                "id": org.id,
                "name": org.name,
                "org_type": org.org_type.value if org.org_type else None,
                "email": org.email,
                "is_active": org.is_active
            },
            "stats": {
                "total_tickets": total_tickets,
                "open_tickets": open_tickets,
                "total_users": total_users,
                "total_devices": total_devices,
                "total_products": total_products,
                "total_sla_policies": total_sla_policies,
                "active_integrations": active_integrations
            },
            "subscription": subscription_data
        }
    except Exception as e:
        import traceback
        raise HTTPException(
            status_code=500,
            detail=f"Error loading dashboard: {str(e)}\n{traceback.format_exc()}"
        )


# Product Catalog Management
@router.get("/products")
async def list_products(
    category: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List all products for the organization"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(Product).filter(Product.organization_id == current_user.organization_id)
        
        if category:
            query = query.filter(Product.category == category)
        
        products = query.all()
        
        return [
            {
                "id": p.id,
                "name": p.name,
                "category": p.category.value if p.category else None,
                "brand": p.brand,
                "description": p.description,
                "default_warranty_months": p.default_warranty_months,
                "extended_warranty_available": p.extended_warranty_available,
                "specifications": p.specifications,
                "common_failures": p.common_failures,
                "recommended_parts": p.recommended_parts,
                "is_active": p.is_active,
                "models_count": len(p.models) if hasattr(p, 'models') else 0
            }
            for p in products
        ]
    except Exception:
        # Table doesn't exist yet
        return []


@router.post("/products")
async def create_product(
    product_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        # Handle specifications if it's a string (from JSON textarea)
        specs = product_data.get("specifications", {})
        if isinstance(specs, str):
            try:
                specs = json.loads(specs) if specs else {}
            except:
                specs = {}
        
        # Handle common_failures - ensure it's a list
        failures = product_data.get("common_failures", [])
        if isinstance(failures, str):
            failures = [f.strip() for f in failures.split('\n') if f.strip()]
        elif not isinstance(failures, list):
            failures = []
        
        # Handle recommended_parts - ensure it's a list
        parts = product_data.get("recommended_parts", [])
        if isinstance(parts, str):
            # If comma-separated string, split it
            parts = [p.strip() for p in parts.split(',') if p.strip()]
        elif not isinstance(parts, list):
            parts = []
        
        # Create the product
        product = Product(
            organization_id=current_user.organization_id,
            name=product_data.get("name"),
            category=product_data.get("category"),
            brand=product_data.get("brand"),
            description=product_data.get("description"),
            default_warranty_months=product_data.get("default_warranty_months", 12),
            extended_warranty_available=product_data.get("extended_warranty_available", False),
            specifications=specs,
            common_failures=failures,
            recommended_parts=parts,
            is_active=product_data.get("is_active", True)
        )
        
        db.add(product)
        db.flush()  # Flush to get the product ID
        
        # Create ProductModel if model_number is provided
        model_number = product_data.get("model_number")
        if model_number:
            from app.models.product import ProductModel
            product_model = ProductModel(
                product_id=product.id,
                organization_id=current_user.organization_id,
                model_number=model_number,
                model_name=product_data.get("name", ""),
                is_active=True
            )
            db.add(product_model)
        
        db.commit()
        db.refresh(product)
        
        return {
            "id": product.id,
            "name": product.name,
            "message": "Product created successfully"
        }
    except Exception as e:
        db.rollback()
        import traceback
        raise HTTPException(status_code=500, detail=f"Error creating product: {str(e)}\n{traceback.format_exc()}")


@router.post("/products/bulk-upload", status_code=status.HTTP_200_OK)
async def bulk_upload_products(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Bulk upload products from Excel file
    
    Excel file should have the following columns:
    - product_name (required): Product name e.g., "Split AC 1.5T"
    - category (required): Product category (ac, refrigerator, washing_machine, tv, microwave, air_purifier, water_purifier, other)
    - brand (optional): Brand name
    - description (optional): Product description
    - default_warranty_months (optional): Default warranty in months (default: 12)
    - extended_warranty_available (optional): true/false (default: false)
    - model_number (optional): Model number for the product model
    - model_name (optional): Model name
    - specifications (optional): JSON string for additional specifications
    - common_failures (optional): Comma-separated list of common failures
    - recommended_parts (optional): Comma-separated list of part IDs
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed. Please install it with: pip install openpyxl"
        )
    
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    # Validate file type
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )
    
    try:
        # Read file content
        contents = await file.read()
        
        # Load workbook
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        worksheet = workbook.active
        
        # Get headers (first row)
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        # Find column indices
        product_name_col = None
        category_col = None
        brand_col = None
        description_col = None
        warranty_col = None
        extended_warranty_col = None
        model_number_col = None
        model_name_col = None
        specifications_col = None
        common_failures_col = None
        recommended_parts_col = None
        
        for idx, header in enumerate(headers):
            if 'product_name' in header or 'name' in header:
                product_name_col = idx + 1
            elif 'category' in header:
                category_col = idx + 1
            elif 'brand' in header:
                brand_col = idx + 1
            elif 'description' in header:
                description_col = idx + 1
            elif 'warranty' in header:
                warranty_col = idx + 1
            elif 'extended_warranty' in header:
                extended_warranty_col = idx + 1
            elif 'model_number' in header:
                model_number_col = idx + 1
            elif 'model_name' in header:
                model_name_col = idx + 1
            elif 'specifications' in header or 'specs' in header:
                specifications_col = idx + 1
            elif 'common_failures' in header or 'failures' in header:
                common_failures_col = idx + 1
            elif 'recommended_parts' in header or 'parts' in header:
                recommended_parts_col = idx + 1
        
        if not product_name_col or not category_col:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain columns: product_name, category"
            )
        
        # Process rows
        successful = 0
        failed = 0
        errors = []
        products_created = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
            # Skip empty rows
            if not any(cell.value for cell in row):
                continue
            
            try:
                # Extract values
                product_name = str(row[product_name_col - 1].value).strip() if row[product_name_col - 1].value else None
                category_str = str(row[category_col - 1].value).strip().lower() if row[category_col - 1].value else None
                
                if not product_name or not category_str:
                    errors.append(f"Row {row_idx}: Missing required fields (product_name, category)")
                    failed += 1
                    continue
                
                # Validate and convert category
                category_map = {
                    'ac': ProductCategory.AC,
                    'refrigerator': ProductCategory.REFRIGERATOR,
                    'washing_machine': ProductCategory.WASHING_MACHINE,
                    'washingmachine': ProductCategory.WASHING_MACHINE,
                    'tv': ProductCategory.TV,
                    'television': ProductCategory.TV,
                    'microwave': ProductCategory.MICROWAVE,
                    'air_purifier': ProductCategory.AIR_PURIFIER,
                    'airpurifier': ProductCategory.AIR_PURIFIER,
                    'water_purifier': ProductCategory.WATER_PURIFIER,
                    'waterpurifier': ProductCategory.WATER_PURIFIER,
                    'other': ProductCategory.OTHER
                }
                
                category = category_map.get(category_str)
                if not category:
                    errors.append(f"Row {row_idx}: Invalid category '{category_str}'. Valid categories: {', '.join(category_map.keys())}")
                    failed += 1
                    continue
                
                # Extract optional fields
                brand = str(row[brand_col - 1].value).strip() if brand_col and row[brand_col - 1].value else None
                description = str(row[description_col - 1].value).strip() if description_col and row[description_col - 1].value else None
                
                default_warranty = 12
                if warranty_col and row[warranty_col - 1].value:
                    try:
                        default_warranty = int(row[warranty_col - 1].value)
                    except:
                        pass
                
                extended_warranty = False
                if extended_warranty_col and row[extended_warranty_col - 1].value:
                    ext_val = str(row[extended_warranty_col - 1].value).strip().lower()
                    extended_warranty = ext_val in ['true', 'yes', '1', 'y']
                
                model_number = str(row[model_number_col - 1].value).strip() if model_number_col and row[model_number_col - 1].value else None
                model_name = str(row[model_name_col - 1].value).strip() if model_name_col and row[model_name_col - 1].value else None
                
                # Parse specifications (JSON)
                specifications = {}
                if specifications_col and row[specifications_col - 1].value:
                    try:
                        specs_str = str(row[specifications_col - 1].value).strip()
                        specifications = json.loads(specs_str) if specs_str else {}
                    except:
                        errors.append(f"Row {row_idx}: Invalid specifications JSON format")
                
                # Parse common failures (comma-separated)
                common_failures = []
                if common_failures_col and row[common_failures_col - 1].value:
                    failures_str = str(row[common_failures_col - 1].value).strip()
                    common_failures = [f.strip() for f in failures_str.split(',') if f.strip()]
                
                # Parse recommended parts (comma-separated)
                recommended_parts = []
                if recommended_parts_col and row[recommended_parts_col - 1].value:
                    parts_str = str(row[recommended_parts_col - 1].value).strip()
                    recommended_parts = [p.strip() for p in parts_str.split(',') if p.strip()]
                
                # Check if product with same name and category already exists
                existing_product = db.query(Product).filter(
                    Product.organization_id == current_user.organization_id,
                    Product.name == product_name,
                    Product.category == category
                ).first()
                
                if existing_product:
                    # Use existing product
                    product = existing_product
                    products_created.append({
                        "product_id": product.id,
                        "product_name": product.name,
                        "status": "existing"
                    })
                else:
                    # Create new product
                    product = Product(
                        organization_id=current_user.organization_id,
                        name=product_name,
                        category=category,
                        brand=brand,
                        description=description,
                        default_warranty_months=default_warranty,
                        extended_warranty_available=extended_warranty,
                        specifications=specifications,
                        common_failures=common_failures,
                        recommended_parts=recommended_parts,
                        is_active=True
                    )
                    db.add(product)
                    db.flush()
                    
                    products_created.append({
                        "product_id": product.id,
                        "product_name": product.name,
                        "status": "created"
                    })
                
                # Create ProductModel if model_number is provided
                if model_number:
                    # Check if model already exists for this product
                    existing_model = db.query(ProductModel).filter(
                        ProductModel.product_id == product.id,
                        ProductModel.model_number == model_number,
                        ProductModel.organization_id == current_user.organization_id
                    ).first()
                    
                    if not existing_model:
                        product_model = ProductModel(
                            product_id=product.id,
                            organization_id=current_user.organization_id,
                            model_number=model_number,
                            model_name=model_name or product_name,
                            is_active=True
                        )
                        db.add(product_model)
                
                successful += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                failed += 1
                continue
        
        # Commit all successful creations
        if successful > 0:
            db.commit()
        
        return {
            "total": successful + failed,
            "successful": successful,
            "failed": failed,
            "errors": errors[:50] if len(errors) > 50 else errors,  # Limit errors to 50
            "products": products_created[:20],  # Show first 20 created products
            "message": f"Successfully processed {successful} product(s)"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error processing bulk product upload: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )


@router.get("/products/bulk-upload-template")
async def download_bulk_upload_template(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Download Excel template for bulk product upload"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed"
        )
    
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products Template"
        
        # Add headers
        headers = [
            "product_name",
            "category",
            "brand",
            "description",
            "default_warranty_months",
            "extended_warranty_available",
            "model_number",
            "model_name",
            "specifications",
            "common_failures",
            "recommended_parts"
        ]
        
        ws.append(headers)
        
        # Add sample rows
        sample_rows = [
            [
                "Split AC 1.5T",
                "ac",
                "CoolAir",
                "High efficiency split air conditioner",
                "12",
                "true",
                "AC-1.5T-2024",
                "CoolAir Split AC 1.5T 2024",
                '{"capacity": "1.5T", "voltage": "220V", "power": "1500W"}',
                "Compressor failure, Refrigerant leak",
                ""
            ],
            [
                "Front Load Washing Machine",
                "washing_machine",
                "CleanTech",
                "Fully automatic front load washing machine",
                "24",
                "true",
                "WM-FL-2024",
                "CleanTech Front Load WM 2024",
                '{"capacity": "8kg", "energy_rating": "5 star"}',
                "Motor failure, Drain pump issue",
                ""
            ]
        ]
        
        for row in sample_rows:
            ws.append(row)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Get the bytes before creating response
        file_content = output.getvalue()
        output.close()
        
        from fastapi.responses import Response
        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="Product_Bulk_Upload_Template.xlsx"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating template: {str(e)}"
        )


@router.get("/products/{product_id}")
async def get_product(
    product_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        return {
            "id": product.id,
            "name": product.name,
            "category": product.category.value if product.category else None,
            "brand": product.brand,
            "description": product.description,
            "default_warranty_months": product.default_warranty_months,
            "specifications": product.specifications,
            "common_failures": product.common_failures,
            "recommended_parts": product.recommended_parts,
            "is_active": product.is_active,
            "models_count": len(product.models) if hasattr(product, 'models') else 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching product: {str(e)}")


@router.put("/products/{product_id}")
async def update_product(
    product_id: int,
    product_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Update fields
        if "name" in product_data:
            product.name = product_data["name"]
        if "category" in product_data:
            product.category = product_data["category"]
        if "brand" in product_data:
            product.brand = product_data["brand"]
        if "description" in product_data:
            product.description = product_data["description"]
        if "default_warranty_months" in product_data:
            product.default_warranty_months = product_data["default_warranty_months"]
        if "extended_warranty_available" in product_data:
            product.extended_warranty_available = product_data["extended_warranty_available"]
        
        # Handle specifications
        if "specifications" in product_data:
            specs = product_data["specifications"]
            if isinstance(specs, str):
                try:
                    specs = json.loads(specs) if specs.strip() else {}
                except:
                    specs = {}
            product.specifications = specs
        
        # Handle common_failures
        if "common_failures" in product_data:
            failures = product_data["common_failures"]
            if isinstance(failures, str):
                failures = [f.strip() for f in failures.split('\n') if f.strip()]
            product.common_failures = failures if isinstance(failures, list) else []
        
        # Handle recommended_parts
        if "recommended_parts" in product_data:
            parts = product_data["recommended_parts"]
            if isinstance(parts, str):
                parts = [p.strip() for p in parts.split(',') if p.strip()]
            product.recommended_parts = parts if isinstance(parts, list) else []
        
        if "is_active" in product_data:
            product.is_active = product_data["is_active"]
        
        db.commit()
        db.refresh(product)
        
        return {
            "id": product.id,
            "name": product.name,
            "message": "Product updated successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating product: {str(e)}")


@router.delete("/products/{product_id}")
async def delete_product(
    product_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        db.delete(product)
        db.commit()
        
        return {"message": "Product deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting product: {str(e)}")


@router.get("/products/{product_id}/models")
async def list_product_models(
    product_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List models for a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        return [
            {
                "id": m.id,
                "model_number": m.model_number,
                "model_name": m.model_name,
                "is_active": m.is_active
            }
            for m in product.models
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching product models: {str(e)}")


# SLA Policy Management
@router.get("/sla-policies")
async def list_sla_policies(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List all SLA policies for the organization"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policies = db.query(SLAPolicy).filter(
            SLAPolicy.organization_id == current_user.organization_id
        ).all()
        
        return [
            {
                "id": p.id,
                "sla_type": p.sla_type.value if p.sla_type else None,
                "target_hours": p.target_hours,
                "product_category": p.product_category,
                "is_active": p.is_active
            }
            for p in policies
        ]
    except Exception:
        # Table doesn't exist yet
        return []


@router.post("/sla-policies")
async def create_sla_policy(
    policy_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new SLA policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = SLAPolicy(
            organization_id=current_user.organization_id,
            sla_type=policy_data.get("sla_type"),
            target_hours=policy_data.get("target_hours"),
            product_category=policy_data.get("product_category"),
            product_id=policy_data.get("product_id"),
            country_id=policy_data.get("country_id"),
            state_id=policy_data.get("state_id"),
            city_id=policy_data.get("city_id"),
            priority_overrides=policy_data.get("priority_overrides", {}),
            business_hours_only=policy_data.get("business_hours_only", False),
            business_hours=policy_data.get("business_hours", {})
        )
        
        db.add(policy)
        db.commit()
        db.refresh(policy)
        
        return {
            "id": policy.id,
            "message": "SLA policy created successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating SLA policy: {str(e)}")


@router.put("/sla-policies/{policy_id}")
async def update_sla_policy(
    policy_id: int,
    policy_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update an SLA policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = db.query(SLAPolicy).filter(
            SLAPolicy.id == policy_id,
            SLAPolicy.organization_id == current_user.organization_id
        ).first()
        
        if not policy:
            raise HTTPException(status_code=404, detail="SLA policy not found")
        
        # Update fields
        if "sla_type" in policy_data:
            policy.sla_type = policy_data["sla_type"]
        if "target_hours" in policy_data:
            policy.target_hours = policy_data["target_hours"]
        if "product_category" in policy_data:
            policy.product_category = policy_data["product_category"]
        if "priority_overrides" in policy_data:
            policy.priority_overrides = policy_data["priority_overrides"]
        if "business_hours_only" in policy_data:
            policy.business_hours_only = policy_data["business_hours_only"]
        if "business_hours" in policy_data:
            policy.business_hours = policy_data["business_hours"]
        if "is_active" in policy_data:
            policy.is_active = policy_data["is_active"]
        
        db.commit()
        db.refresh(policy)
        
        return {
            "id": policy.id,
            "message": "SLA policy updated successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating SLA policy: {str(e)}")


@router.delete("/sla-policies/{policy_id}")
async def delete_sla_policy(
    policy_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete an SLA policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = db.query(SLAPolicy).filter(
            SLAPolicy.id == policy_id,
            SLAPolicy.organization_id == current_user.organization_id
        ).first()
        
        if not policy:
            raise HTTPException(status_code=404, detail="SLA policy not found")
        
        db.delete(policy)
        db.commit()
        
        return {"message": "SLA policy deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting SLA policy: {str(e)}")


# Service Policy Management
@router.get("/service-policies")
async def list_service_policies(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List all service policies"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policies = db.query(ServicePolicy).filter(
            ServicePolicy.organization_id == current_user.organization_id
        ).all()
        
        return [
            {
                "id": p.id,
                "policy_type": p.policy_type,
                "rules": p.rules,
                "product_category": p.product_category,
                "product_id": p.product_id,
                "country_id": p.country_id,
                "state_id": p.state_id,
                "city_id": p.city_id,
                "is_active": p.is_active,
                "created_at": p.created_at.isoformat() if p.created_at else None,
                "updated_at": p.updated_at.isoformat() if p.updated_at else None
            }
            for p in policies
        ]
    except Exception:
        # Table doesn't exist yet
        return []


@router.get("/service-policies/{policy_id}")
async def get_service_policy(
    policy_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single service policy by ID"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = db.query(ServicePolicy).filter(
            ServicePolicy.id == policy_id,
            ServicePolicy.organization_id == current_user.organization_id
        ).first()
        
        if not policy:
            raise HTTPException(status_code=404, detail="Service policy not found")
        
        return {
            "id": policy.id,
            "policy_type": policy.policy_type,
            "rules": policy.rules,
            "product_category": policy.product_category,
            "product_id": policy.product_id,
            "country_id": policy.country_id,
            "state_id": policy.state_id,
            "city_id": policy.city_id,
            "is_active": policy.is_active,
            "created_at": policy.created_at.isoformat() if policy.created_at else None,
            "updated_at": policy.updated_at.isoformat() if policy.updated_at else None
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching service policy: {str(e)}")


@router.post("/service-policies")
async def create_service_policy(
    policy_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new service policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = ServicePolicy(
            organization_id=current_user.organization_id,
            policy_type=policy_data.get("policy_type"),
            rules=policy_data.get("rules", {}),
            product_category=policy_data.get("product_category"),
            product_id=policy_data.get("product_id"),
            country_id=policy_data.get("country_id"),
            state_id=policy_data.get("state_id"),
            city_id=policy_data.get("city_id"),
            is_active=policy_data.get("is_active", True)
        )
        
        db.add(policy)
        db.commit()
        db.refresh(policy)
        
        return {
            "id": policy.id,
            "policy_type": policy.policy_type,
            "rules": policy.rules,
            "product_category": policy.product_category,
            "product_id": policy.product_id,
            "country_id": policy.country_id,
            "state_id": policy.state_id,
            "city_id": policy.city_id,
            "is_active": policy.is_active,
            "message": "Service policy created successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating service policy: {str(e)}")


@router.put("/service-policies/{policy_id}")
async def update_service_policy(
    policy_id: int,
    policy_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update a service policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = db.query(ServicePolicy).filter(
            ServicePolicy.id == policy_id,
            ServicePolicy.organization_id == current_user.organization_id
        ).first()
        
        if not policy:
            raise HTTPException(status_code=404, detail="Service policy not found")
        
        # Update fields
        if "policy_type" in policy_data:
            policy.policy_type = policy_data["policy_type"]
        if "rules" in policy_data:
            policy.rules = policy_data["rules"]
        if "product_category" in policy_data:
            policy.product_category = policy_data["product_category"]
        if "product_id" in policy_data:
            policy.product_id = policy_data["product_id"]
        if "country_id" in policy_data:
            policy.country_id = policy_data["country_id"]
        if "state_id" in policy_data:
            policy.state_id = policy_data["state_id"]
        if "city_id" in policy_data:
            policy.city_id = policy_data["city_id"]
        if "is_active" in policy_data:
            policy.is_active = policy_data["is_active"]
        
        db.commit()
        db.refresh(policy)
        
        return {
            "id": policy.id,
            "policy_type": policy.policy_type,
            "rules": policy.rules,
            "product_category": policy.product_category,
            "product_id": policy.product_id,
            "country_id": policy.country_id,
            "state_id": policy.state_id,
            "city_id": policy.city_id,
            "is_active": policy.is_active,
            "message": "Service policy updated successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating service policy: {str(e)}")


@router.delete("/service-policies/{policy_id}")
async def delete_service_policy(
    policy_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete a service policy"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        policy = db.query(ServicePolicy).filter(
            ServicePolicy.id == policy_id,
            ServicePolicy.organization_id == current_user.organization_id
        ).first()
        
        if not policy:
            raise HTTPException(status_code=404, detail="Service policy not found")
        
        db.delete(policy)
        db.commit()
        
        return {"message": "Service policy deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting service policy: {str(e)}")


# Integration Management
@router.get("/integrations")
async def list_integrations(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List all integrations for the organization"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        integrations = db.query(Integration).filter(
            Integration.organization_id == current_user.organization_id
        ).all()
        
        return [
            {
                "id": i.id,
                "name": i.name,
                "integration_type": i.integration_type.value if i.integration_type else None,
                "provider": i.provider,
                "status": i.status.value if i.status else None,
                "is_active": i.is_active,
                "last_sync_at": i.last_sync_at.isoformat() if i.last_sync_at else None,
                "last_sync_stats": (i.config or {}).get("last_sync_stats")
            }
            for i in integrations
        ]
    except Exception:
        # Table doesn't exist yet
        return []


@router.post("/integrations")
async def create_integration(
    integration_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new integration"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        integration = Integration(
            organization_id=current_user.organization_id,
            name=integration_data.get("name"),
            integration_type=integration_data.get("integration_type"),
            provider=integration_data.get("provider"),
            config=integration_data.get("config", {}),
            webhook_url=integration_data.get("webhook_url"),
            api_endpoint=integration_data.get("api_endpoint"),
            sync_direction=integration_data.get("sync_direction", "bidirectional"),
            sync_frequency=integration_data.get("sync_frequency", "realtime")
        )
        
        db.add(integration)
        db.commit()
        db.refresh(integration)
        
        return {
            "id": integration.id,
            "message": "Integration created successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating integration: {str(e)}")


@router.put("/integrations/{integration_id}")
async def update_integration(
    integration_id: int,
    integration_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update an integration"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        integration = db.query(Integration).filter(
            Integration.id == integration_id,
            Integration.organization_id == current_user.organization_id
        ).first()
        
        if not integration:
            raise HTTPException(status_code=404, detail="Integration not found")
        
        # Update fields
        if "name" in integration_data:
            integration.name = integration_data["name"]
        if "integration_type" in integration_data:
            integration.integration_type = integration_data["integration_type"]
        if "provider" in integration_data:
            integration.provider = integration_data["provider"]
        if "config" in integration_data:
            integration.config = integration_data["config"]
        if "webhook_url" in integration_data:
            integration.webhook_url = integration_data["webhook_url"]
        if "api_endpoint" in integration_data:
            integration.api_endpoint = integration_data["api_endpoint"]
        if "sync_direction" in integration_data:
            integration.sync_direction = integration_data["sync_direction"]
        if "sync_frequency" in integration_data:
            integration.sync_frequency = integration_data["sync_frequency"]
        if "is_active" in integration_data:
            integration.is_active = integration_data["is_active"]
        
        db.commit()
        db.refresh(integration)
        
        return {
            "id": integration.id,
            "message": "Integration updated successfully"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating integration: {str(e)}")


@router.delete("/integrations/{integration_id}")
async def delete_integration(
    integration_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete an integration"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        integration = db.query(Integration).filter(
            Integration.id == integration_id,
            Integration.organization_id == current_user.organization_id
        ).first()
        
        if not integration:
            raise HTTPException(status_code=404, detail="Integration not found")
        
        db.delete(integration)
        db.commit()
        
        return {"message": "Integration deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting integration: {str(e)}")


# Partner Management (for OEM)
@router.get("/partners")
async def list_partners(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List service partners (for OEM organizations)"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    org = db.query(Organization).filter(Organization.id == current_user.organization_id).first()
    if not org or org.org_type.value != "oem":
        return []
    
    # Get child organizations (service partners)
    partners = db.query(Organization).filter(
        Organization.parent_organization_id == org.id
    ).all()
    
    return [
        {
            "id": p.id,
            "name": p.name,
            "email": p.email,
            "phone": p.phone,
            "is_active": p.is_active
        }
        for p in partners
    ]


@router.post("/partners")
async def create_partner(
    partner_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create/register a service partner (for OEM)"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    org = db.query(Organization).filter(Organization.id == current_user.organization_id).first()
    if not org or org.org_type.value != "oem":
        raise HTTPException(status_code=400, detail="Only OEM organizations can create partners")
    
    # Create service partner organization
    partner = Organization(
        name=partner_data.get("name"),
        org_type="service_company",
        email=partner_data.get("email"),
        phone=partner_data.get("phone"),
        address=partner_data.get("address"),
        parent_organization_id=org.id,
        country_id=partner_data.get("country_id"),
        state_id=partner_data.get("state_id"),
        city_id=partner_data.get("city_id")
    )
    
    db.add(partner)
    db.commit()
    db.refresh(partner)
    
    # Create organization hierarchy
    hierarchy = OrganizationHierarchy(
        oem_organization_id=org.id,
        service_partner_id=partner.id,
        product_categories=partner_data.get("product_categories", []),
        service_regions=partner_data.get("service_regions", [])
    )
    
    db.add(hierarchy)
    db.commit()
    
    return {
        "id": partner.id,
        "message": "Service partner created successfully"
    }


@router.get("/ai/cost-to-serve")
async def get_cost_to_serve(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Estimate cost-to-serve per model and city"""
    tickets = db.query(Ticket).filter(
        Ticket.organization_id == current_user.organization_id,
        Ticket.created_at.isnot(None)
    ).all()

    part_costs = {p.id: (p.cost_price or p.selling_price or 0) for p in db.query(Part).all()}
    groups = {}
    for t in tickets:
        if not t.device_id:
            continue
        device = db.query(Device).filter(Device.id == t.device_id).first()
        if not device:
            continue
        key = (device.model_number or "unknown", t.city_id or 0)
        cost = 300
        if t.parts_used:
            for item in t.parts_used:
                if isinstance(item, dict):
                    pid = item.get("part_id")
                    qty = item.get("quantity", 1)
                    cost += (part_costs.get(pid, 0) * qty)
        groups.setdefault(key, []).append(cost)

    results = []
    for (model, city_id), costs in groups.items():
        avg_cost = sum(costs) / len(costs) if costs else 0
        city = db.query(City).filter(City.id == city_id).first() if city_id else None
        results.append({
            "model_number": model,
            "city_id": city_id,
            "city_name": city.name if city else "N/A",
            "avg_cost": round(avg_cost, 2),
            "tickets": len(costs)
        })

    return results


@router.get("/ai/inventory-forecast")
async def get_inventory_forecast(
    days: int = 30,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Inventory forecast by city and part"""
    since = datetime.now(timezone.utc) - timedelta(days=30)
    transactions = db.query(InventoryTransaction).join(Inventory).filter(
        Inventory.organization_id == current_user.organization_id,
        InventoryTransaction.transaction_type == "out",
        InventoryTransaction.created_at >= since
    ).all()

    usage = {}
    for tx in transactions:
        key = (tx.part_id, tx.inventory_id)
        usage[key] = usage.get(key, 0) + tx.quantity

    forecasts = []
    for (part_id, inventory_id), qty in usage.items():
        inv = db.query(Inventory).filter(Inventory.id == inventory_id).first()
        part = db.query(Part).filter(Part.id == part_id).first()
        avg_per_day = qty / 30 if qty else 0
        city = db.query(City).filter(City.id == inv.city_id).first() if inv else None
        forecasts.append({
            "part_id": part_id,
            "part_name": part.name if part else f"Part {part_id}",
            "city_id": inv.city_id if inv else None,
            "city_name": city.name if city else "N/A",
            "forecast_days": days,
            "predicted_demand": round(avg_per_day * days, 2)
        })

    return forecasts[:50]


@router.post("/ai/route-optimizer")
async def optimize_routes_for_engineer(
    payload: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Optimize routes for an engineer's assigned tickets"""
    engineer_id = payload.get("engineer_id")
    if not engineer_id:
        raise HTTPException(status_code=400, detail="engineer_id is required")

    engineer = db.query(User).filter(
        User.id == engineer_id,
        User.organization_id == current_user.organization_id
    ).first()
    if not engineer:
        raise HTTPException(status_code=404, detail="Engineer not found")

    tickets = db.query(Ticket).filter(
        Ticket.assigned_engineer_id == engineer.id,
        Ticket.status.in_([TicketStatus.ASSIGNED, TicketStatus.IN_PROGRESS])
    ).all()

    ticket_locations = {}
    for t in tickets:
        if t.service_latitude and t.service_longitude:
            ticket_locations[t.id] = (float(t.service_latitude), float(t.service_longitude))

    if not ticket_locations:
        return {"optimized_order": [], "message": "No ticket locations found"}

    if engineer.current_location_lat and engineer.current_location_lng:
        engineer_location = (float(engineer.current_location_lat), float(engineer.current_location_lng))
    else:
        engineer_location = (0.0, 0.0)

    result = await route_optimizer.optimize_routes(
        engineer_id=engineer.id,
        ticket_ids=list(ticket_locations.keys()),
        engineer_location=engineer_location,
        ticket_locations=ticket_locations
    )
    return result


@router.post("/subscription/upgrade")
async def upgrade_subscription(
    upgrade_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Upgrade organization subscription to a new plan"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    plan_id = upgrade_data.get("plan_id")
    billing_period_str = upgrade_data.get("billing_period", "monthly")
    
    if not plan_id:
        raise HTTPException(status_code=400, detail="plan_id is required")
    
    # Get the plan
    plan = db.query(Plan).filter(Plan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    if not plan.is_active:
        raise HTTPException(status_code=400, detail="Plan is not active")
    
    # Get current subscription
    current_subscription = db.query(Subscription).filter(
        Subscription.organization_id == current_user.organization_id
    ).first()
    
    # Calculate price based on billing period
    try:
        billing_period = BillingPeriod(billing_period_str)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid billing_period. Must be 'monthly' or 'annual'")
    
    subscription_price = plan.monthly_price if billing_period == BillingPeriod.MONTHLY else plan.annual_price
    
    if not subscription_price or subscription_price <= 0:
        raise HTTPException(status_code=400, detail="Plan price is invalid")
    
    # Calculate dates
    from datetime import datetime, timedelta, timezone
    start_date = datetime.now(timezone.utc)
    
    if billing_period == BillingPeriod.MONTHLY:
        end_date = start_date + timedelta(days=30)
    else:  # annual
        end_date = start_date + timedelta(days=365)
    
    if current_subscription:
        # Update existing subscription
        current_subscription.plan_id = plan.id
        current_subscription.billing_period = billing_period
        current_subscription.current_price = float(subscription_price)
        current_subscription.start_date = start_date
        current_subscription.end_date = end_date
        current_subscription.status = "active"
        
        db.commit()
        db.refresh(current_subscription)
        
        return {
            "id": current_subscription.id,
            "plan_name": plan.name,
            "billing_period": billing_period.value,
            "price": float(subscription_price),
            "message": "Subscription upgraded successfully"
        }
    else:
        # Create new subscription
        subscription = Subscription(
            organization_id=current_user.organization_id,
            plan_id=plan.id,
            billing_period=billing_period,
            current_price=float(subscription_price),
            currency="INR",
            status="active",
            start_date=start_date,
            end_date=end_date
        )
        
        db.add(subscription)
        db.commit()
        db.refresh(subscription)
        
        return {
            "id": subscription.id,
            "plan_name": plan.name,
            "billing_period": billing_period.value,
            "price": float(subscription_price),
            "message": "Subscription created successfully"
        }


# Analytics & KPIs
@router.get("/analytics")
async def get_org_analytics(
    period: str = "30d",
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get organization-wide analytics"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    org_id = current_user.organization_id
    
    # Calculate period
    from datetime import datetime, timedelta, timezone
    if period == "7d":
        start_date = datetime.now(timezone.utc) - timedelta(days=7)
    elif period == "30d":
        start_date = datetime.now(timezone.utc) - timedelta(days=30)
    elif period == "90d":
        start_date = datetime.now(timezone.utc) - timedelta(days=90)
    elif period == "1y":
        start_date = datetime.now(timezone.utc) - timedelta(days=365)
    else:
        start_date = None
    
    try:
        # Ticket analytics
        ticket_query = db.query(Ticket).filter(Ticket.organization_id == org_id)
        if start_date:
            ticket_query = ticket_query.filter(Ticket.created_at >= start_date)
        
        total_tickets = ticket_query.count()
        resolved_tickets = ticket_query.filter(Ticket.status == TicketStatus.RESOLVED).count()
        open_tickets = ticket_query.filter(Ticket.status.in_([TicketStatus.CREATED, TicketStatus.ASSIGNED, TicketStatus.IN_PROGRESS])).count()
        closed_tickets = ticket_query.filter(Ticket.status == TicketStatus.CLOSED).count()
        
        # Calculate SLA compliance (tickets resolved within SLA)
        # For now, we'll use resolved/total as a simple metric
        sla_compliance = (resolved_tickets / total_tickets * 100) if total_tickets > 0 else 0
        
        # Get device count (handle if table doesn't exist)
        try:
            device_count = db.query(func.count(Device.id)).filter(Device.organization_id == org_id).scalar() or 0
        except Exception:
            device_count = 0
        
        # Get user count
        try:
            user_count = db.query(func.count(User.id)).filter(User.organization_id == org_id).scalar() or 0
        except Exception:
            user_count = 0
        
        # Get product count
        try:
            from app.models.product import Product
            product_count = db.query(func.count(Product.id)).filter(Product.organization_id == org_id).scalar() or 0
        except Exception:
            product_count = 0
        
        # Get parts count
        try:
            parts_count = db.query(func.count(Part.id)).filter(Part.is_active == True).scalar() or 0
        except Exception:
            parts_count = 0
        
        # Generate daily/weekly trends for charts
        daily_trends = []
        if start_date:
            from datetime import timedelta
            current_date = start_date
            end_date = datetime.now(timezone.utc)
            
            # Determine interval based on period length
            days_diff = (end_date - start_date).days
            if days_diff <= 30:
                # Daily data for short periods
                interval_days = 1
                date_format = "%b %d"
            elif days_diff <= 90:
                # Weekly data for medium periods
                interval_days = 7
                date_format = "%b %d"
            else:
                # Monthly data for long periods
                interval_days = 30
                date_format = "%b %Y"
            
            while current_date <= end_date:
                period_start = current_date.replace(hour=0, minute=0, second=0, microsecond=0)
                period_end = period_start + timedelta(days=interval_days)
                if period_end > end_date:
                    period_end = end_date
                
                period_tickets = db.query(func.count(Ticket.id)).filter(
                    Ticket.organization_id == org_id,
                    Ticket.created_at >= period_start,
                    Ticket.created_at < period_end
                ).scalar() or 0
                
                period_resolved = db.query(func.count(Ticket.id)).filter(
                    Ticket.organization_id == org_id,
                    Ticket.status == TicketStatus.RESOLVED,
                    Ticket.updated_at >= period_start,
                    Ticket.updated_at < period_end
                ).scalar() or 0
                
                daily_trends.append({
                    "date": period_start.strftime("%Y-%m-%d"),
                    "day": period_start.strftime(date_format),
                    "tickets": period_tickets,
                    "resolved": period_resolved
                })
                
                current_date = period_end
        else:
            # For "all time", show monthly data for last 12 months
            from datetime import timedelta
            end_date = datetime.now(timezone.utc)
            for i in range(12):
                month_start = (end_date - timedelta(days=30 * (12 - i))).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                month_end = month_start + timedelta(days=30)
                
                month_tickets = db.query(func.count(Ticket.id)).filter(
                    Ticket.organization_id == org_id,
                    Ticket.created_at >= month_start,
                    Ticket.created_at < month_end
                ).scalar() or 0
                
                month_resolved = db.query(func.count(Ticket.id)).filter(
                    Ticket.organization_id == org_id,
                    Ticket.status == TicketStatus.RESOLVED,
                    Ticket.updated_at >= month_start,
                    Ticket.updated_at < month_end
                ).scalar() or 0
                
                daily_trends.append({
                    "date": month_start.strftime("%Y-%m-%d"),
                    "day": month_start.strftime("%b %Y"),
                    "tickets": month_tickets,
                    "resolved": month_resolved
                })
        
        # Generate status distribution
        status_distribution = {}
        try:
            status_counts = db.query(
                Ticket.status,
                func.count(Ticket.id).label('count')
            ).filter(
                Ticket.organization_id == org_id
            )
            if start_date:
                status_counts = status_counts.filter(Ticket.created_at >= start_date)
            status_counts = status_counts.group_by(Ticket.status).all()
            
            for status, count in status_counts:
                status_distribution[status.value if hasattr(status, 'value') else str(status)] = count
        except Exception:
            status_distribution = {}
        
        return {
            "period": period,
            "tickets": {
                "total": total_tickets,
                "resolved": resolved_tickets,
                "open": open_tickets,
                "closed": closed_tickets,
                "sla_compliance": round(sla_compliance, 2)
            },
            "users": {
                "total": user_count
            },
            "devices": {
                "total": device_count
            },
            "products": {
                "total": product_count
            },
            "parts": {
                "total": parts_count
            },
            "daily_trends": daily_trends,
            "status_distribution": status_distribution
        }
    except Exception as e:
        import traceback
        print(f"Error loading analytics: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Error loading analytics: {str(e)}")


# Inventory Management
@router.get("/inventory/parts")
async def list_parts(
    search: Optional[str] = None,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List all parts with optional search"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(Part).filter(Part.is_active == True)
        
        # Add search filter if provided
        if search:
            query = query.filter(
                or_(
                    Part.name.ilike(f"%{search}%"),
                    Part.sku.ilike(f"%{search}%"),
                    Part.description.ilike(f"%{search}%")
                )
            )
        
        parts = query.order_by(Part.name).all()
        
        return [
            {
                "id": p.id,
                "sku": p.sku,
                "name": p.name,
                "description": p.description,
                "cost_price": float(p.cost_price) if p.cost_price else None,
                "selling_price": float(p.selling_price) if p.selling_price else None,
                "unit": p.unit,
                "applicable_products": p.applicable_products or [],
                "compatible_models": p.compatible_models or [],
                "is_active": p.is_active
            }
            for p in parts
        ]
    except Exception as e:
        import traceback
        from sqlalchemy.exc import ProgrammingError
        
        error_msg = str(e)
        error_trace = traceback.format_exc()
        
        # Log the error
        print(f"Error fetching parts: {error_msg}")
        print(error_trace)
        
        # Check if it's a table not found error
        if isinstance(e, ProgrammingError) or "doesn't exist" in error_msg.lower() or "Table" in error_msg:
            # Table doesn't exist yet, return empty array
            return []
        
        # For other errors, raise HTTPException with details
        raise HTTPException(
            status_code=500, 
            detail=f"Error fetching parts: {error_msg}"
        )


# IMPORTANT: These specific routes must come BEFORE the parameterized route /inventory/parts/{part_id}
# to avoid FastAPI trying to match "bulk-upload-template" as an integer part_id
@router.get("/inventory/parts/bulk-upload-template")
async def download_parts_bulk_upload_template(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Download Excel template for bulk part upload"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed"
        )
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Parts Template"
        
        headers = [
            "sku",
            "name",
            "description",
            "cost_price",
            "selling_price",
            "unit",
            "applicable_products",
            "compatible_models"
        ]
        
        ws.append(headers)
        
        # Add sample rows
        sample_rows = [
            [
                "AC-COMP-001",
                "AC Compressor",
                "1.5T AC Compressor",
                "5000.00",
                "7500.00",
                "piece",
                "ac",
                "AC-1.5T-2024, AC-2T-2024"
            ],
            [
                "WM-MOTOR-001",
                "Washing Machine Motor",
                "Front load washing machine motor",
                "3000.00",
                "4500.00",
                "piece",
                "washing_machine",
                "WM-FL-2024"
            ]
        ]
        
        for row in sample_rows:
            ws.append(row)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.getvalue()
        output.close()
        
        from fastapi.responses import Response
        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="Parts_Bulk_Upload_Template.xlsx"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating template: {str(e)}"
        )


@router.post("/inventory/parts/bulk-upload", status_code=status.HTTP_200_OK)
async def bulk_upload_parts(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Bulk upload parts from Excel file
    
    Excel file should have the following columns:
    - sku (required): Unique SKU code
    - name (required): Part name
    - description (optional): Part description
    - cost_price (optional): Cost price
    - selling_price (optional): Selling price
    - unit (optional): Unit of measurement (default: piece)
    - applicable_products (optional): Comma-separated product categories
    - compatible_models (optional): Comma-separated model numbers
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed"
        )
    
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    # Validate file type
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        worksheet = workbook.active
        
        # Get headers
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        # Find column indices
        sku_col = None
        name_col = None
        description_col = None
        cost_price_col = None
        selling_price_col = None
        unit_col = None
        applicable_products_col = None
        compatible_models_col = None
        
        for idx, header in enumerate(headers):
            if 'sku' in header:
                sku_col = idx + 1
            elif 'name' in header and 'model' not in header:
                name_col = idx + 1
            elif 'description' in header:
                description_col = idx + 1
            elif 'cost_price' in header or 'cost' in header:
                cost_price_col = idx + 1
            elif 'selling_price' in header or 'selling' in header:
                selling_price_col = idx + 1
            elif 'unit' in header:
                unit_col = idx + 1
            elif 'applicable_products' in header or 'applicable' in header:
                applicable_products_col = idx + 1
            elif 'compatible_models' in header or 'compatible' in header:
                compatible_models_col = idx + 1
        
        if not sku_col or not name_col:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain columns: sku, name"
            )
        
        # Process rows
        successful = 0
        failed = 0
        errors = []
        parts_created = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):
                continue
            
            try:
                sku = str(row[sku_col - 1].value).strip() if row[sku_col - 1].value else None
                name = str(row[name_col - 1].value).strip() if row[name_col - 1].value else None
                
                if not sku or not name:
                    errors.append(f"Row {row_idx}: Missing required fields (sku, name)")
                    failed += 1
                    continue
                
                # Check if SKU already exists
                existing = db.query(Part).filter(Part.sku == sku).first()
                if existing:
                    errors.append(f"Row {row_idx}: SKU '{sku}' already exists")
                    failed += 1
                    continue
                
                # Extract optional fields
                description = str(row[description_col - 1].value).strip() if description_col and row[description_col - 1].value else None
                
                cost_price = None
                if cost_price_col and row[cost_price_col - 1].value:
                    try:
                        cost_price = float(row[cost_price_col - 1].value)
                    except:
                        pass
                
                selling_price = None
                if selling_price_col and row[selling_price_col - 1].value:
                    try:
                        selling_price = float(row[selling_price_col - 1].value)
                    except:
                        pass
                
                unit = "piece"
                if unit_col and row[unit_col - 1].value:
                    unit = str(row[unit_col - 1].value).strip()
                
                applicable_products = []
                if applicable_products_col and row[applicable_products_col - 1].value:
                    products_str = str(row[applicable_products_col - 1].value).strip()
                    applicable_products = [p.strip() for p in products_str.split(',') if p.strip()]
                
                compatible_models = []
                if compatible_models_col and row[compatible_models_col - 1].value:
                    models_str = str(row[compatible_models_col - 1].value).strip()
                    compatible_models = [m.strip() for m in models_str.split(',') if m.strip()]
                
                # Create part
                part = Part(
                    sku=sku,
                    name=name,
                    description=description,
                    cost_price=cost_price,
                    selling_price=selling_price,
                    unit=unit,
                    applicable_products=applicable_products,
                    compatible_models=compatible_models,
                    is_active=True
                )
                
                db.add(part)
                parts_created.append({
                    "part_id": None,  # Will be set after flush
                    "sku": sku,
                    "name": name,
                    "status": "created"
                })
                successful += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                failed += 1
                continue
        
        # Commit all successful creations
        if successful > 0:
            db.commit()
            # Update part IDs
            for i, part_data in enumerate(parts_created):
                part = db.query(Part).filter(Part.sku == part_data["sku"]).first()
                if part:
                    parts_created[i]["part_id"] = part.id
        
        return {
            "total": successful + failed,
            "successful": successful,
            "failed": failed,
            "errors": errors[:50] if len(errors) > 50 else errors,
            "parts": parts_created[:20],
            "message": f"Successfully processed {successful} part(s)"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error processing bulk part upload: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )


@router.get("/inventory/parts/{part_id}")
async def get_part(
    part_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single part by ID"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        part = db.query(Part).filter(Part.id == part_id).first()
        
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        
        # Get inventory stock for this part
        inventory = db.query(Inventory).filter(
            Inventory.part_id == part_id,
            Inventory.organization_id == current_user.organization_id
        ).all()
        
        # Get products that use this part
        product_parts = db.query(ProductPart).options(
            joinedload(ProductPart.product)
        ).filter(
            ProductPart.part_id == part_id,
            ProductPart.organization_id == current_user.organization_id
        ).all()
        
        return {
            "id": part.id,
            "sku": part.sku,
            "name": part.name,
            "description": part.description,
            "cost_price": part.cost_price,
            "selling_price": part.selling_price,
            "unit": part.unit,
            "applicable_products": part.applicable_products,
            "compatible_models": part.compatible_models,
            "is_active": part.is_active,
            "inventory": [
                {
                    "id": inv.id,
                    "current_stock": inv.current_stock,
                    "min_threshold": inv.min_threshold,
                    "max_threshold": inv.max_threshold,
                    "is_low_stock": inv.is_low_stock,
                    "warehouse_name": inv.warehouse_name,
                    "city_id": inv.city_id,
                    "state_id": inv.state_id
                }
                for inv in inventory
            ],
            "used_by_products": [
                {
                    "product_id": pp.product_id,
                    "product_name": pp.product.name if pp.product else None,
                    "is_required": pp.is_required,
                    "is_common": pp.is_common
                }
                for pp in product_parts
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching part: {str(e)}")


@router.post("/inventory/parts")
async def create_part(
    part_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new part"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        # Check if SKU already exists
        existing = db.query(Part).filter(Part.sku == part_data.get("sku")).first()
        if existing:
            raise HTTPException(status_code=400, detail="SKU already exists")
        
        part = Part(
            sku=part_data.get("sku"),
            name=part_data.get("name"),
            description=part_data.get("description"),
            cost_price=part_data.get("cost_price"),
            selling_price=part_data.get("selling_price"),
            unit=part_data.get("unit", "piece"),
            applicable_products=part_data.get("applicable_products", []),
            compatible_models=part_data.get("compatible_models", []),
            is_active=True
        )
        
        db.add(part)
        db.commit()
        db.refresh(part)
        
        return {
            "id": part.id,
            "sku": part.sku,
            "message": "Part created successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating part: {str(e)}")


@router.put("/inventory/parts/{part_id}")
async def update_part(
    part_id: int,
    part_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update a part"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        part = db.query(Part).filter(Part.id == part_id).first()
        
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        
        # Check SKU uniqueness if SKU is being changed
        if "sku" in part_data and part_data["sku"] != part.sku:
            existing = db.query(Part).filter(Part.sku == part_data["sku"]).first()
            if existing:
                raise HTTPException(status_code=400, detail="SKU already exists")
        
        # Update fields
        if "sku" in part_data:
            part.sku = part_data["sku"]
        if "name" in part_data:
            part.name = part_data["name"]
        if "description" in part_data:
            part.description = part_data["description"]
        if "cost_price" in part_data:
            part.cost_price = part_data["cost_price"]
        if "selling_price" in part_data:
            part.selling_price = part_data["selling_price"]
        if "unit" in part_data:
            part.unit = part_data["unit"]
        if "applicable_products" in part_data:
            part.applicable_products = part_data["applicable_products"]
        if "compatible_models" in part_data:
            part.compatible_models = part_data["compatible_models"]
        if "is_active" in part_data:
            part.is_active = part_data["is_active"]
        
        db.commit()
        db.refresh(part)
        
        return {
            "id": part.id,
            "sku": part.sku,
            "message": "Part updated successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating part: {str(e)}")


@router.delete("/inventory/parts/{part_id}")
async def delete_part(
    part_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete a part (soft delete by setting is_active=False)"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        part = db.query(Part).filter(Part.id == part_id).first()
        
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        
        # Check if part is used in inventory
        inventory_count = db.query(func.count(Inventory.id)).filter(
            Inventory.part_id == part_id,
            Inventory.organization_id == current_user.organization_id
        ).scalar() or 0
        
        if inventory_count > 0:
            # Soft delete - just deactivate
            part.is_active = False
            db.commit()
            return {"message": "Part deactivated successfully (has inventory entries)"}
        else:
            # Hard delete if no inventory
            db.delete(part)
            db.commit()
            return {"message": "Part deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting part: {str(e)}")


@router.get("/inventory/stock")
async def get_inventory(
    city_id: Optional[int] = None,
    state_id: Optional[int] = None,
    part_id: Optional[int] = None,
    low_stock_only: bool = False,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get inventory levels"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(Inventory).options(joinedload(Inventory.part)).filter(
            Inventory.organization_id == current_user.organization_id
        )
        
        if city_id:
            query = query.filter(Inventory.city_id == city_id)
        if state_id:
            query = query.filter(Inventory.state_id == state_id)
        if part_id:
            query = query.filter(Inventory.part_id == part_id)
        if low_stock_only:
            query = query.filter(Inventory.is_low_stock == True)
        
        inventory_items = query.all()
        
        return [
            {
                "id": inv.id,
                "part_id": inv.part_id,
                "part_name": inv.part.name if inv.part else None,
                "sku": inv.part.sku if inv.part else None,
                "current_stock": inv.current_stock,
                "min_threshold": inv.min_threshold,
                "max_threshold": inv.max_threshold,
                "reserved_stock": inv.reserved_stock,
                "is_low_stock": inv.is_low_stock,
                "warehouse_name": inv.warehouse_name,
                "city_id": inv.city_id,
                "state_id": inv.state_id,
                "country_id": inv.country_id,
                "last_restocked_at": inv.last_restocked_at.isoformat() if inv.last_restocked_at else None
            }
            for inv in inventory_items
        ]
    except Exception:
        return []


@router.get("/inventory/stock/{inventory_id}")
async def get_inventory_entry(
    inventory_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single inventory entry by ID"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        inventory = db.query(Inventory).options(
            joinedload(Inventory.part)
        ).filter(
            Inventory.id == inventory_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not inventory:
            raise HTTPException(status_code=404, detail="Inventory entry not found")
        
        # Get products that use this part
        product_parts = db.query(ProductPart).options(
            joinedload(ProductPart.product)
        ).filter(
            ProductPart.part_id == inventory.part_id,
            ProductPart.organization_id == current_user.organization_id
        ).all()
        
        return {
            "id": inventory.id,
            "part_id": inventory.part_id,
            "part_name": inventory.part.name if inventory.part else None,
            "sku": inventory.part.sku if inventory.part else None,
            "current_stock": inventory.current_stock,
            "min_threshold": inventory.min_threshold,
            "max_threshold": inventory.max_threshold,
            "reserved_stock": inventory.reserved_stock,
            "is_low_stock": inventory.is_low_stock,
            "warehouse_name": inventory.warehouse_name,
            "country_id": inventory.country_id,
            "state_id": inventory.state_id,
            "city_id": inventory.city_id,
            "last_restocked_at": inventory.last_restocked_at.isoformat() if inventory.last_restocked_at else None,
            "created_at": inventory.created_at.isoformat() if inventory.created_at else None,
            "updated_at": inventory.updated_at.isoformat() if inventory.updated_at else None,
            "used_by_products": [
                {
                    "product_id": pp.product_id,
                    "product_name": pp.product.name if pp.product else None,
                    "is_required": pp.is_required,
                    "is_common": pp.is_common
                }
                for pp in product_parts
            ]
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching inventory: {str(e)}")


@router.post("/inventory/stock")
async def create_inventory(
    inventory_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create or update inventory entry"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        part_id = inventory_data.get("part_id")
        if not part_id:
            raise HTTPException(status_code=400, detail="part_id is required")
        
        # Helper function to convert empty strings to None and ensure int type
        def clean_id(value):
            # Handle None, empty string, string 'None', or whitespace-only strings
            if value is None:
                return None
            if isinstance(value, str):
                value = value.strip()
                if value == '' or value.lower() == 'none' or value == 'null':
                    return None
            # Try to convert to int
            try:
                if value:
                    return int(value)
                return None
            except (ValueError, TypeError):
                return None
        
        # Clean location IDs - explicitly handle empty strings
        country_id_raw = inventory_data.get("country_id")
        state_id_raw = inventory_data.get("state_id")
        city_id_raw = inventory_data.get("city_id")
        
        country_id = clean_id(country_id_raw)
        state_id = clean_id(state_id_raw)
        city_id = clean_id(city_id_raw)

        if not city_id:
            raise HTTPException(status_code=400, detail="city_id is required")

        city = db.query(City).filter(City.id == city_id).first()
        if not city:
            raise HTTPException(status_code=404, detail="City not found")
        state = db.query(State).filter(State.id == city.state_id).first()
        if not state:
            raise HTTPException(status_code=404, detail="State not found for city")
        if state_id and state_id != city.state_id:
            raise HTTPException(status_code=400, detail="state_id does not match city_id")
        state_id = city.state_id
        country_id = state.country_id
        
        # Ensure part_id is an integer
        try:
            part_id = int(part_id) if part_id else None
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="Invalid part_id")
        
        # Check if inventory entry already exists
        existing = db.query(Inventory).filter(
            Inventory.part_id == part_id,
            Inventory.organization_id == current_user.organization_id,
            Inventory.city_id == city_id,
            Inventory.state_id == state_id
        ).first()
        
        if existing:
            # Update existing
            existing.current_stock = inventory_data.get("current_stock", existing.current_stock)
            existing.min_threshold = inventory_data.get("min_threshold", existing.min_threshold)
            existing.max_threshold = inventory_data.get("max_threshold")
            existing.warehouse_name = inventory_data.get("warehouse_name")
            existing.is_low_stock = existing.current_stock <= existing.min_threshold
            
            db.commit()
            db.refresh(existing)
            
            return {
                "id": existing.id,
                "message": "Inventory updated successfully"
            }
        else:
            # Create new - ensure all values are properly typed
            max_threshold_raw = inventory_data.get("max_threshold")
            max_threshold = None
            if max_threshold_raw is not None and max_threshold_raw != '':
                try:
                    max_threshold = int(max_threshold_raw)
                except (ValueError, TypeError):
                    max_threshold = None
            
            inventory = Inventory(
                part_id=part_id,  # Already converted to int above
                organization_id=current_user.organization_id,
                country_id=country_id,
                state_id=state_id,
                city_id=city_id,
                warehouse_name=inventory_data.get("warehouse_name") or None,
                current_stock=int(inventory_data.get("current_stock", 0) or 0),
                min_threshold=int(inventory_data.get("min_threshold", 0) or 0),
                max_threshold=max_threshold,
                reserved_stock=0,
                is_low_stock=False
            )
            
            inventory.is_low_stock = inventory.current_stock <= inventory.min_threshold
            
            db.add(inventory)
            db.commit()
            db.refresh(inventory)
            
            return {
                "id": inventory.id,
                "message": "Inventory created successfully"
            }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error managing inventory: {str(e)}")


@router.post("/inventory/stock/bulk-upload", status_code=status.HTTP_200_OK)
async def bulk_upload_inventory(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Bulk upload inventory/stock from Excel file
    
    Excel file should have the following columns:
    - part_sku (required): Part SKU code (must exist)
    - current_stock (required): Current stock quantity
    - min_threshold (required): Minimum stock threshold
    - max_threshold (optional): Maximum stock threshold
    - warehouse_name (optional): Warehouse name
    - country_id (optional): Country ID
    - state_id (optional): State ID
    - city_id (optional): City ID
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed"
        )
    
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    # Validate file type
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file type. Please upload an Excel file (.xlsx or .xls)"
        )
    
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        worksheet = workbook.active
        
        # Get headers
        headers = []
        for cell in worksheet[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        # Find column indices
        part_sku_col = None
        current_stock_col = None
        min_threshold_col = None
        max_threshold_col = None
        warehouse_name_col = None
        country_id_col = None
        state_id_col = None
        city_id_col = None
        
        for idx, header in enumerate(headers):
            if 'part_sku' in header or ('sku' in header and 'part' in header):
                part_sku_col = idx + 1
            elif 'current_stock' in header or 'stock' in header:
                current_stock_col = idx + 1
            elif 'min_threshold' in header or 'min' in header:
                min_threshold_col = idx + 1
            elif 'max_threshold' in header or 'max' in header:
                max_threshold_col = idx + 1
            elif 'warehouse' in header:
                warehouse_name_col = idx + 1
            elif 'country_id' in header:
                country_id_col = idx + 1
            elif 'state_id' in header:
                state_id_col = idx + 1
            elif 'city_id' in header:
                city_id_col = idx + 1
        
        if not all([part_sku_col, current_stock_col, min_threshold_col]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain columns: part_sku, current_stock, min_threshold"
            )
        
        # Process rows
        successful = 0
        failed = 0
        errors = []
        inventory_created = []
        
        def clean_id(value):
            if value is None:
                return None
            if isinstance(value, str):
                value = value.strip()
                if value == '' or value.lower() == 'none' or value == 'null':
                    return None
            try:
                if value:
                    return int(value)
                return None
            except (ValueError, TypeError):
                return None
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
            if not any(cell.value for cell in row):
                continue
            
            try:
                part_sku = str(row[part_sku_col - 1].value).strip() if row[part_sku_col - 1].value else None
                current_stock_val = row[current_stock_col - 1].value if row[current_stock_col - 1].value else None
                min_threshold_val = row[min_threshold_col - 1].value if row[min_threshold_col - 1].value else None
                
                if not part_sku:
                    errors.append(f"Row {row_idx}: Missing part_sku")
                    failed += 1
                    continue
                
                # Find part by SKU
                part = db.query(Part).filter(Part.sku == part_sku).first()
                if not part:
                    errors.append(f"Row {row_idx}: Part with SKU '{part_sku}' not found")
                    failed += 1
                    continue
                
                # Parse stock values
                try:
                    current_stock = int(current_stock_val) if current_stock_val is not None else 0
                except:
                    current_stock = 0
                
                try:
                    min_threshold = int(min_threshold_val) if min_threshold_val is not None else 0
                except:
                    min_threshold = 0
                
                max_threshold = None
                if max_threshold_col and row[max_threshold_col - 1].value:
                    try:
                        max_threshold = int(row[max_threshold_col - 1].value)
                    except:
                        pass
                
                warehouse_name = None
                if warehouse_name_col and row[warehouse_name_col - 1].value:
                    warehouse_name = str(row[warehouse_name_col - 1].value).strip()
                
                country_id = None
                if country_id_col and row[country_id_col - 1].value:
                    country_id = clean_id(row[country_id_col - 1].value)
                
                state_id = None
                if state_id_col and row[state_id_col - 1].value:
                    state_id = clean_id(row[state_id_col - 1].value)
                
                city_id = None
                if city_id_col and row[city_id_col - 1].value:
                    city_id = clean_id(row[city_id_col - 1].value)

                if not city_id:
                    errors.append(f"Row {row_idx}: city_id is required")
                    failed += 1
                    continue

                city = db.query(City).filter(City.id == city_id).first()
                if not city:
                    errors.append(f"Row {row_idx}: City {city_id} not found")
                    failed += 1
                    continue
                state = db.query(State).filter(State.id == city.state_id).first()
                if not state:
                    errors.append(f"Row {row_idx}: State not found for city {city_id}")
                    failed += 1
                    continue
                if state_id and state_id != city.state_id:
                    errors.append(f"Row {row_idx}: state_id does not match city_id")
                    failed += 1
                    continue
                state_id = city.state_id
                country_id = state.country_id
                
                # Check if inventory entry already exists
                existing = db.query(Inventory).filter(
                    Inventory.part_id == part.id,
                    Inventory.organization_id == current_user.organization_id,
                    Inventory.city_id == city_id,
                    Inventory.state_id == state_id
                ).first()
                
                if existing:
                    # Update existing
                    existing.current_stock = current_stock
                    existing.min_threshold = min_threshold
                    existing.max_threshold = max_threshold
                    existing.warehouse_name = warehouse_name
                    existing.is_low_stock = current_stock <= min_threshold
                    inventory_created.append({
                        "inventory_id": existing.id,
                        "part_sku": part_sku,
                        "status": "updated"
                    })
                else:
                    # Create new
                    inventory = Inventory(
                        part_id=part.id,
                        organization_id=current_user.organization_id,
                        country_id=country_id,
                        state_id=state_id,
                        city_id=city_id,
                        warehouse_name=warehouse_name,
                        current_stock=current_stock,
                        min_threshold=min_threshold,
                        max_threshold=max_threshold,
                        reserved_stock=0,
                        is_low_stock=current_stock <= min_threshold
                    )
                    db.add(inventory)
                    db.flush()
                    inventory_created.append({
                        "inventory_id": inventory.id,
                        "part_sku": part_sku,
                        "status": "created"
                    })
                
                successful += 1
                
            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
                failed += 1
                continue
        
        # Commit all successful creations
        if successful > 0:
            db.commit()
        
        return {
            "total": successful + failed,
            "successful": successful,
            "failed": failed,
            "errors": errors[:50] if len(errors) > 50 else errors,
            "inventory": inventory_created[:20],
            "message": f"Successfully processed {successful} inventory entry/entries"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error processing bulk inventory upload: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error processing Excel file: {str(e)}"
        )


@router.get("/inventory/stock/bulk-upload-template")
async def download_inventory_bulk_upload_template(
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Download Excel template for bulk inventory/stock upload"""
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel processing library (openpyxl) is not installed"
        )
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Template"
        
        headers = [
            "part_sku",
            "current_stock",
            "min_threshold",
            "max_threshold",
            "warehouse_name",
            "country_id",
            "state_id",
            "city_id"
        ]
        
        ws.append(headers)
        
        # Add sample rows
        sample_rows = [
            [
                "AC-COMP-001",
                "50",
                "10",
                "100",
                "Main Warehouse",
                "",
                "",
                ""
            ],
            [
                "WM-MOTOR-001",
                "30",
                "5",
                "50",
                "City Warehouse",
                "",
                "",
                ""
            ]
        ]
        
        for row in sample_rows:
            ws.append(row)
        
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content = output.getvalue()
        output.close()
        
        from fastapi.responses import Response
        return Response(
            content=file_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="Inventory_Stock_Bulk_Upload_Template.xlsx"',
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error generating template: {str(e)}"
        )


@router.put("/inventory/stock/{inventory_id}")
async def update_inventory(
    inventory_id: int,
    inventory_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Update an inventory entry"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        inventory = db.query(Inventory).filter(
            Inventory.id == inventory_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not inventory:
            raise HTTPException(status_code=404, detail="Inventory entry not found")
        
        # Helper function to convert empty strings to None and ensure int type
        def clean_id(value):
            if value is None or value == '' or value == 'None':
                return None
            try:
                return int(value) if value else None
            except (ValueError, TypeError):
                return None
        
        # Update fields
        if "current_stock" in inventory_data:
            inventory.current_stock = int(inventory_data["current_stock"])
        if "min_threshold" in inventory_data:
            inventory.min_threshold = int(inventory_data["min_threshold"])
        if "max_threshold" in inventory_data:
            max_val = inventory_data["max_threshold"]
            inventory.max_threshold = int(max_val) if max_val and max_val != '' else None
        if "warehouse_name" in inventory_data:
            inventory.warehouse_name = inventory_data["warehouse_name"]
        if "city_id" in inventory_data:
            city_id = clean_id(inventory_data["city_id"])
            if not city_id:
                raise HTTPException(status_code=400, detail="city_id is required")
            city = db.query(City).filter(City.id == city_id).first()
            if not city:
                raise HTTPException(status_code=404, detail="City not found")
            state = db.query(State).filter(State.id == city.state_id).first()
            if not state:
                raise HTTPException(status_code=404, detail="State not found for city")
            inventory.city_id = city_id
            inventory.state_id = city.state_id
            inventory.country_id = state.country_id
        elif "state_id" in inventory_data or "country_id" in inventory_data:
            raise HTTPException(status_code=400, detail="Update city_id to change location")
        
        # Recalculate low stock status
        inventory.is_low_stock = inventory.current_stock <= inventory.min_threshold
        
        db.commit()
        db.refresh(inventory)
        
        return {
            "id": inventory.id,
            "current_stock": inventory.current_stock,
            "message": "Inventory updated successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating inventory: {str(e)}")


@router.delete("/inventory/stock/{inventory_id}")
async def delete_inventory(
    inventory_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Delete an inventory entry"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        inventory = db.query(Inventory).filter(
            Inventory.id == inventory_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not inventory:
            raise HTTPException(status_code=404, detail="Inventory entry not found")
        
        # Check if there are transactions
        transaction_count = db.query(func.count(InventoryTransaction.id)).filter(
            InventoryTransaction.inventory_id == inventory_id
        ).scalar() or 0
        
        if transaction_count > 0:
            raise HTTPException(
                status_code=400, 
                detail="Cannot delete inventory entry with transaction history. Set stock to 0 instead."
            )
        
        db.delete(inventory)
        db.commit()
        
        return {"message": "Inventory entry deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting inventory: {str(e)}")


@router.post("/inventory/stock/{inventory_id}/adjust")
async def adjust_stock(
    inventory_id: int,
    adjustment_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Adjust inventory stock (add or remove)"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        inventory = db.query(Inventory).filter(
            Inventory.id == inventory_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not inventory:
            raise HTTPException(status_code=404, detail="Inventory not found")
        
        transaction_type = adjustment_data.get("transaction_type", "adjustment")  # in, out, adjustment
        quantity = adjustment_data.get("quantity", 0)
        notes = adjustment_data.get("notes", "")
        
        if quantity == 0:
            raise HTTPException(status_code=400, detail="Quantity cannot be zero")
        
        previous_stock = inventory.current_stock
        
        if transaction_type == "in":
            new_stock = previous_stock + abs(quantity)
        elif transaction_type == "out":
            new_stock = max(0, previous_stock - abs(quantity))
        else:  # adjustment
            new_stock = quantity
        
        inventory.current_stock = new_stock
        inventory.is_low_stock = new_stock <= inventory.min_threshold
        
        if transaction_type == "in":
            from datetime import datetime, timezone
            inventory.last_restocked_at = datetime.now(timezone.utc)
        
        # Create transaction log
        transaction = InventoryTransaction(
            part_id=inventory.part_id,
            inventory_id=inventory.id,
            transaction_type=transaction_type,
            quantity=abs(quantity) if transaction_type != "adjustment" else (new_stock - previous_stock),
            previous_stock=previous_stock,
            new_stock=new_stock,
            performed_by_id=current_user.id,
            notes=notes
        )
        
        db.add(transaction)
        db.commit()
        db.refresh(inventory)
        
        return {
            "id": inventory.id,
            "current_stock": inventory.current_stock,
            "message": "Stock adjusted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error adjusting stock: {str(e)}")


@router.get("/inventory/transactions")
async def list_transactions(
    part_id: Optional[int] = None,
    inventory_id: Optional[int] = None,
    transaction_type: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List inventory transactions with filters"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(InventoryTransaction).join(Inventory).filter(
            Inventory.organization_id == current_user.organization_id
        )
        
        if part_id:
            query = query.filter(InventoryTransaction.part_id == part_id)
        if inventory_id:
            query = query.filter(InventoryTransaction.inventory_id == inventory_id)
        if transaction_type:
            query = query.filter(InventoryTransaction.transaction_type == transaction_type)
        
        total_count = query.count()
        transactions = query.order_by(InventoryTransaction.created_at.desc()).offset(offset).limit(limit).all()
        
        return {
            "total": total_count,
            "limit": limit,
            "offset": offset,
            "transactions": [
                {
                    "id": t.id,
                    "part_id": t.part_id,
                    "part_name": t.part.name if t.part else None,
                    "sku": t.part.sku if t.part else None,
                    "inventory_id": t.inventory_id,
                    "transaction_type": t.transaction_type,
                    "quantity": t.quantity,
                    "previous_stock": t.previous_stock,
                    "new_stock": t.new_stock,
                    "notes": t.notes,
                    "ticket_id": t.ticket_id,
                    "performed_by": t.performed_by.full_name if t.performed_by else None,
                    "performed_by_id": t.performed_by_id,
                    "created_at": t.created_at.isoformat()
                }
                for t in transactions
            ]
        }
    except Exception:
        return {"total": 0, "limit": limit, "offset": offset, "transactions": []}


@router.get("/inventory/transactions/{transaction_id}")
async def get_transaction(
    transaction_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single transaction by ID"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        transaction = db.query(InventoryTransaction).options(
            joinedload(InventoryTransaction.part),
            joinedload(InventoryTransaction.inventory),
            joinedload(InventoryTransaction.performed_by)
        ).join(Inventory).filter(
            InventoryTransaction.id == transaction_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        return {
            "id": transaction.id,
            "part_id": transaction.part_id,
            "part_name": transaction.part.name if transaction.part else None,
            "sku": transaction.part.sku if transaction.part else None,
            "inventory_id": transaction.inventory_id,
            "inventory_warehouse": transaction.inventory.warehouse_name if transaction.inventory else None,
            "transaction_type": transaction.transaction_type,
            "quantity": transaction.quantity,
            "previous_stock": transaction.previous_stock,
            "new_stock": transaction.new_stock,
            "notes": transaction.notes,
            "ticket_id": transaction.ticket_id,
            "performed_by": {
                "id": transaction.performed_by.id if transaction.performed_by else None,
                "name": transaction.performed_by.full_name if transaction.performed_by else None,
                "email": transaction.performed_by.email if transaction.performed_by else None
            },
            "created_at": transaction.created_at.isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching transaction: {str(e)}")


@router.get("/inventory/reorder-requests")
async def list_reorder_requests(
    status: Optional[str] = None,
    part_id: Optional[int] = None,
    limit: int = 50,
    offset: int = 0,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """List reorder requests with filters"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(ReorderRequest).options(joinedload(ReorderRequest.part)).filter(
            ReorderRequest.organization_id == current_user.organization_id
        )
        
        if status:
            query = query.filter(ReorderRequest.status == status)
        if part_id:
            query = query.filter(ReorderRequest.part_id == part_id)
        
        total_count = query.count()
        requests = query.order_by(ReorderRequest.created_at.desc()).offset(offset).limit(limit).all()
        
        return {
            "total": total_count,
            "limit": limit,
            "offset": offset,
            "requests": [
                {
                    "id": req.id,
                    "part_id": req.part_id,
                    "part_name": req.part.name if req.part else None,
                    "sku": req.part.sku if req.part else None,
                    "inventory_id": req.inventory_id,
                    "requested_quantity": req.requested_quantity,
                    "current_stock": req.current_stock,
                    "min_threshold": req.min_threshold,
                    "status": req.status,
                    "requested_by": {
                        "id": req.requested_by.id if req.requested_by else None,
                        "name": req.requested_by.full_name if req.requested_by else None
                    },
                    "approved_by": {
                        "id": req.approved_by.id if req.approved_by else None,
                        "name": req.approved_by.full_name if req.approved_by else None
                    },
                    "created_at": req.created_at.isoformat(),
                    "approved_at": req.approved_at.isoformat() if req.approved_at else None,
                    "updated_at": req.updated_at.isoformat() if req.updated_at else None
                }
                for req in requests
            ]
        }
    except Exception:
        return {"total": 0, "limit": limit, "offset": offset, "requests": []}


@router.get("/inventory/reorder-requests/{request_id}")
async def get_reorder_request(
    request_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get a single reorder request by ID"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        request = db.query(ReorderRequest).options(
            joinedload(ReorderRequest.part),
            joinedload(ReorderRequest.inventory),
            joinedload(ReorderRequest.requested_by),
            joinedload(ReorderRequest.approved_by)
        ).filter(
            ReorderRequest.id == request_id,
            ReorderRequest.organization_id == current_user.organization_id
        ).first()
        
        if not request:
            raise HTTPException(status_code=404, detail="Reorder request not found")
        
        return {
            "id": request.id,
            "part_id": request.part_id,
            "part_name": request.part.name if request.part else None,
            "sku": request.part.sku if request.part else None,
            "inventory_id": request.inventory_id,
            "warehouse_name": request.inventory.warehouse_name if request.inventory else None,
            "requested_quantity": request.requested_quantity,
            "current_stock": request.current_stock,
            "min_threshold": request.min_threshold,
            "status": request.status,
            "requested_by": {
                "id": request.requested_by.id if request.requested_by else None,
                "name": request.requested_by.full_name if request.requested_by else None,
                "email": request.requested_by.email if request.requested_by else None
            },
            "approved_by": {
                "id": request.approved_by.id if request.approved_by else None,
                "name": request.approved_by.full_name if request.approved_by else None,
                "email": request.approved_by.email if request.approved_by else None
            },
            "created_at": request.created_at.isoformat(),
            "approved_at": request.approved_at.isoformat() if request.approved_at else None,
            "updated_at": request.updated_at.isoformat() if request.updated_at else None
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching reorder request: {str(e)}")


@router.post("/inventory/reorder-requests")
async def create_reorder_request(
    request_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Create a new reorder request for low stock inventory"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        inventory_id = request_data.get("inventory_id")
        requested_quantity = request_data.get("requested_quantity")
        
        if not inventory_id:
            raise HTTPException(status_code=400, detail="inventory_id is required")
        if not requested_quantity or requested_quantity <= 0:
            raise HTTPException(status_code=400, detail="requested_quantity must be greater than 0")
        
        # Get inventory record
        inventory = db.query(Inventory).filter(
            Inventory.id == inventory_id,
            Inventory.organization_id == current_user.organization_id
        ).first()
        
        if not inventory:
            raise HTTPException(status_code=404, detail="Inventory not found")
        
        # Check if there's already a pending request for this inventory
        existing_request = db.query(ReorderRequest).filter(
            ReorderRequest.inventory_id == inventory_id,
            ReorderRequest.status == "pending"
        ).first()
        
        if existing_request:
            raise HTTPException(status_code=400, detail="A pending reorder request already exists for this inventory")
        
        # Create reorder request
        reorder_request = ReorderRequest(
            part_id=inventory.part_id,
            inventory_id=inventory_id,
            organization_id=current_user.organization_id,
            requested_quantity=requested_quantity,
            current_stock=inventory.current_stock,
            min_threshold=inventory.min_threshold,
            status="pending",
            requested_by_id=current_user.id
        )
        
        db.add(reorder_request)
        db.commit()
        db.refresh(reorder_request)
        
        # Load relationships
        db.refresh(reorder_request)
        reorder_request = db.query(ReorderRequest).options(
            joinedload(ReorderRequest.part),
            joinedload(ReorderRequest.inventory),
            joinedload(ReorderRequest.requested_by)
        ).filter(ReorderRequest.id == reorder_request.id).first()
        
        return {
            "id": reorder_request.id,
            "part_id": reorder_request.part_id,
            "part_name": reorder_request.part.name if reorder_request.part else None,
            "inventory_id": reorder_request.inventory_id,
            "warehouse_name": reorder_request.inventory.warehouse_name if reorder_request.inventory else None,
            "requested_quantity": reorder_request.requested_quantity,
            "current_stock": reorder_request.current_stock,
            "min_threshold": reorder_request.min_threshold,
            "status": reorder_request.status,
            "requested_by": {
                "id": reorder_request.requested_by.id if reorder_request.requested_by else None,
                "name": reorder_request.requested_by.full_name if reorder_request.requested_by else None
            },
            "created_at": reorder_request.created_at.isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating reorder request: {str(e)}")


@router.post("/inventory/reorder-requests/{request_id}/approve")
async def approve_reorder_request(
    request_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Approve a reorder request"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        request = db.query(ReorderRequest).filter(
            ReorderRequest.id == request_id,
            ReorderRequest.organization_id == current_user.organization_id
        ).first()
        
        if not request:
            raise HTTPException(status_code=404, detail="Reorder request not found")
        
        if request.status != "pending":
            raise HTTPException(status_code=400, detail=f"Request is already {request.status}")
        
        request.status = "approved"
        request.approved_by_id = current_user.id
        from datetime import datetime, timezone
        request.approved_at = datetime.now(timezone.utc)
        
        db.commit()
        db.refresh(request)
        
        # Load relationships for response
        request = db.query(ReorderRequest).options(
            joinedload(ReorderRequest.part),
            joinedload(ReorderRequest.inventory),
            joinedload(ReorderRequest.requested_by),
            joinedload(ReorderRequest.approved_by)
        ).filter(ReorderRequest.id == request_id).first()
        
        return {
            "id": request.id,
            "status": request.status,
            "approved_by": {
                "id": request.approved_by.id if request.approved_by else None,
                "name": request.approved_by.full_name if request.approved_by else None
            },
            "approved_at": request.approved_at.isoformat() if request.approved_at else None,
            "message": "Reorder request approved successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error approving reorder request: {str(e)}")


@router.post("/inventory/reorder-requests/{request_id}/reject")
async def reject_reorder_request(
    request_id: int,
    rejection_reason: Optional[str] = Body(None, embed=True),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Reject a reorder request"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        request = db.query(ReorderRequest).filter(
            ReorderRequest.id == request_id,
            ReorderRequest.organization_id == current_user.organization_id
        ).first()
        
        if not request:
            raise HTTPException(status_code=404, detail="Reorder request not found")
        
        if request.status != "pending":
            raise HTTPException(status_code=400, detail=f"Request is already {request.status}")
        
        request.status = "rejected"
        request.approved_by_id = current_user.id
        from datetime import datetime, timezone
        request.approved_at = datetime.now(timezone.utc)
        
        db.commit()
        db.refresh(request)
        
        return {
            "id": request.id,
            "status": request.status,
            "message": "Reorder request rejected",
            "rejection_reason": rejection_reason
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error rejecting reorder request: {str(e)}")


@router.post("/inventory/reorder-requests/{request_id}/fulfill")
async def fulfill_reorder_request(
    request_id: int,
    fulfillment_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Mark a reorder request as fulfilled (after stock is received)"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        request = db.query(ReorderRequest).options(
            joinedload(ReorderRequest.inventory)
        ).filter(
            ReorderRequest.id == request_id,
            ReorderRequest.organization_id == current_user.organization_id
        ).first()
        
        if not request:
            raise HTTPException(status_code=404, detail="Reorder request not found")
        
        if request.status != "approved":
            raise HTTPException(status_code=400, detail="Request must be approved before fulfillment")
        
        # Update inventory stock
        received_quantity = fulfillment_data.get("received_quantity", request.requested_quantity)
        inventory = request.inventory
        
        if inventory:
            inventory.current_stock += received_quantity
            inventory.is_low_stock = inventory.current_stock < inventory.min_threshold
            
            # Create transaction record
            transaction = InventoryTransaction(
                inventory_id=inventory.id,
                part_id=inventory.part_id,
                transaction_type="restock",
                quantity=received_quantity,
                previous_stock=inventory.current_stock - received_quantity,
                new_stock=inventory.current_stock,
                performed_by_id=current_user.id,
                notes=f"Fulfilled reorder request #{request_id}"
            )
            db.add(transaction)
        
        request.status = "fulfilled"
        db.commit()
        
        return {
            "id": request.id,
            "status": request.status,
            "received_quantity": received_quantity,
            "message": "Reorder request fulfilled and stock updated"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error fulfilling reorder request: {str(e)}")


# Product-Part Relationship Management
@router.get("/products/{product_id}/parts")
async def get_product_parts(
    product_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get all parts associated with a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        # Get product
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        # Get product-part relationships
        product_parts = db.query(ProductPart).options(
            joinedload(ProductPart.part)
        ).filter(
            ProductPart.product_id == product_id,
            ProductPart.organization_id == current_user.organization_id
        ).all()
        
        # Get inventory stock for each part
        result = []
        for pp in product_parts:
            # Get inventory stock
            inventory = db.query(Inventory).filter(
                Inventory.part_id == pp.part_id,
                Inventory.organization_id == current_user.organization_id
            ).first()
            
            result.append({
                "id": pp.id,
                "part_id": pp.part_id,
                "part_name": pp.part.name if pp.part else None,
                "sku": pp.part.sku if pp.part else None,
                "cost_price": pp.part.cost_price if pp.part else None,
                "selling_price": pp.part.selling_price if pp.part else None,
                "is_required": pp.is_required,
                "is_common": pp.is_common,
                "usage_frequency": pp.usage_frequency,
                "notes": pp.notes,
                "stock": {
                    "current": inventory.current_stock if inventory else 0,
                    "min_threshold": inventory.min_threshold if inventory else 0,
                    "is_low_stock": inventory.is_low_stock if inventory else False,
                    "warehouse": inventory.warehouse_name if inventory else None
                } if inventory else None
            })
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching product parts: {str(e)}")


@router.post("/products/{product_id}/parts")
async def add_product_part(
    product_id: int,
    part_data: dict = Body(...),
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Add a part to a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        # Verify product belongs to organization
        product = db.query(Product).filter(
            Product.id == product_id,
            Product.organization_id == current_user.organization_id
        ).first()
        
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        
        part_id = part_data.get("part_id")
        if not part_id:
            raise HTTPException(status_code=400, detail="part_id is required")
        
        # Check if relationship already exists
        existing = db.query(ProductPart).filter(
            ProductPart.product_id == product_id,
            ProductPart.part_id == part_id,
            ProductPart.organization_id == current_user.organization_id
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="Part already linked to this product")
        
        # Create relationship
        product_part = ProductPart(
            product_id=product_id,
            part_id=part_id,
            organization_id=current_user.organization_id,
            is_required=part_data.get("is_required", False),
            is_common=part_data.get("is_common", True),
            usage_frequency=part_data.get("usage_frequency", "occasional"),
            notes=part_data.get("notes")
        )
        
        db.add(product_part)
        db.commit()
        db.refresh(product_part)
        
        return {
            "id": product_part.id,
            "message": "Part linked to product successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error linking part to product: {str(e)}")


@router.delete("/products/{product_id}/parts/{part_id}")
async def remove_product_part(
    product_id: int,
    part_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Remove a part from a product"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product_part = db.query(ProductPart).filter(
            ProductPart.product_id == product_id,
            ProductPart.part_id == part_id,
            ProductPart.organization_id == current_user.organization_id
        ).first()
        
        if not product_part:
            raise HTTPException(status_code=404, detail="Product-part relationship not found")
        
        db.delete(product_part)
        db.commit()
        
        return {"message": "Part removed from product successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error removing part from product: {str(e)}")


@router.get("/parts/{part_id}/products")
async def get_part_products(
    part_id: int,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get all products that use a specific part"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        product_parts = db.query(ProductPart).options(
            joinedload(ProductPart.product)
        ).filter(
            ProductPart.part_id == part_id,
            ProductPart.organization_id == current_user.organization_id
        ).all()
        
        return [
            {
                "id": pp.id,
                "product_id": pp.product_id,
                "product_name": pp.product.name if pp.product else None,
                "product_category": pp.product.category.value if pp.product and pp.product.category else None,
                "is_required": pp.is_required,
                "is_common": pp.is_common,
                "usage_frequency": pp.usage_frequency
            }
            for pp in product_parts
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching part products: {str(e)}")


@router.get("/inventory/stock-with-products")
async def get_inventory_with_products(
    low_stock_only: bool = False,
    current_user: User = Depends(require_role([UserRole.ORGANIZATION_ADMIN])),
    db: Session = Depends(get_db)
):
    """Get inventory stock with information about which products use each part"""
    if not current_user.organization_id:
        raise HTTPException(status_code=400, detail="User must be associated with an organization")
    
    try:
        query = db.query(Inventory).options(
            joinedload(Inventory.part)
        ).filter(
            Inventory.organization_id == current_user.organization_id
        )
        
        if low_stock_only:
            query = query.filter(Inventory.is_low_stock == True)
        
        inventory_items = query.all()
        
        result = []
        for inv in inventory_items:
            # Get products that use this part
            product_parts = db.query(ProductPart).options(
                joinedload(ProductPart.product)
            ).filter(
                ProductPart.part_id == inv.part_id,
                ProductPart.organization_id == current_user.organization_id
            ).all()
            
            result.append({
                "id": inv.id,
                "part_id": inv.part_id,
                "part_name": inv.part.name if inv.part else None,
                "sku": inv.part.sku if inv.part else None,
                "current_stock": inv.current_stock,
                "min_threshold": inv.min_threshold,
                "max_threshold": inv.max_threshold,
                "is_low_stock": inv.is_low_stock,
                "warehouse_name": inv.warehouse_name,
                "used_by_products": [
                    {
                        "product_id": pp.product_id,
                        "product_name": pp.product.name if pp.product else None,
                        "is_required": pp.is_required,
                        "is_common": pp.is_common
                    }
                    for pp in product_parts
                ]
            })
        
        return result
    except Exception:
        return []

